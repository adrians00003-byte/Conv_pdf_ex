from pathlib import Path
import shutil
import uuid
import sys
from datetime import datetime, date
from typing import Any

from fastapi import FastAPI, Request, UploadFile, File, HTTPException
from fastapi.templating import Jinja2Templates
from fastapi.responses import FileResponse

ROOT_DIR = Path(__file__).resolve().parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.append(str(ROOT_DIR))

from main_files.pipeline import run_batch


BASE_DIR = Path(__file__).resolve().parent
UPLOAD_BASE = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "output"
ARCHIVE_DIR = BASE_DIR / "archive"
FAILED_DIR = BASE_DIR / "failed"

MAX_FILES = 100
ALLOWED_EXTENSIONS = {".pdf"}

for path in [UPLOAD_BASE, OUTPUT_DIR, ARCHIVE_DIR, FAILED_DIR]:
    path.mkdir(parents=True, exist_ok=True)

app = FastAPI()
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

def sanitize_filename(filename: str) -> str:
  return Path(filename).name

def _format_value_for_preview(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value

def read_excel_preview(excel_path: Path, max_rows: int = 100) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(excel_path, data_only=True)
    previews = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        sheet_rows = []
        
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue

            row_data = {headers[i]: _format_value_for_preview(row[i]) for i in range(len(headers))}            
            sheet_rows.append(row_data)
    
            if len(sheet_rows) >= max_rows:
                break
        previews.append({
            "name": sheet.title,
            "headers": headers,
            "rows": sheet_rows,
        })
    return previews
    
@app.get("/")
def home(request: Request):
    return templates.TemplateResponse(
        request,
          "index.html",
          {
            "title": "Invoice Parser",
            "descripction": "Wgraj fakture PDF, sprawdz dane i wygeneruj Excel",
            "upload_files": None,
            "preview_rows": None,
            "download_url": None,
            "errors": None,
          }
    )
@app.post("/upload")
async def upload(request: Request, invoice_files: list[UploadFile] = File(...)):
  if len(invoice_files) == 0:
        raise HTTPException(status_code=400, detail="Nie wybrano pliku/plikow.")
  if len(invoice_files) > MAX_FILES:
        raise HTTPException(status_code=400, detail=f"Można przesłać maksymalnie {MAX_FILES} plikow.")
  
  session_id = uuid.uuid4().hex
  session_dir = UPLOAD_BASE / session_id
  session_dir.mkdir(parents=True, exist_ok=True)

  saved_files = []
  for upload_file in invoice_files:
    filename = sanitize_filename(upload_file.filename or "")
    if not filename:
        saved_files.append({"filename": "Nieznany", "status": "Nie można przetworzyć pliku bez nazwy."})
        continue
    
    extension = Path(filename).suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        saved_files.append({"filename": filename, "status": f"Nieobsługiwany format pliku: {extension}"})
        continue
    
    destination = session_dir / filename
    counter = 1
    while destination.exists():
        destination = session_dir / f"{Path(filename).stem}_{counter}{extension}"
        counter += 1

    upload_file.file.seek(0)
    with destination.open("wb") as out_file:
        upload_file.file.seek(0)
        shutil.copyfileobj(upload_file.file, out_file)

    saved_files.append({"filename": destination.name, "status": "Plik przesłany pomyślnie."})
  
  output_path = OUTPUT_DIR / f"invoices_{session_id}.xlsx"
  try:
    run_batch(
        inbox_dir=session_dir,
        out_path=output_path,
        archive_dir=ARCHIVE_DIR,
        failed_dir=FAILED_DIR,
    )
  except Exception as e:
      return templates.TemplateResponse(
          request,
            "index.html",
            {
                
                "title": "Invoice Parser",
                "description": "Wgraj faktury PDF, sprawdz dane i wygeneruj Excel",
                "upload_files": saved_files,
                "preview_rows": None,
                "download_url": None,
                "error": f"Błąd przetwarzania: {e}",
            }
        )
  if not output_path.exists():
    raise HTTPException(status_code=400, detail="Nie ma pliku exela.")
  preview_rows = read_excel_preview(output_path)
  return templates.TemplateResponse(
      request,  
        "index.html",
        {
            
            "title": "Invoice Parser",
            "description": "Wgraj faktury PDF, sprawdz dane i wygeneruj Excel",
            "upload_files": saved_files,
            "preview_rows": preview_rows,
            "download_url": f"/download/{output_path.name}",
            "error": None,
        }
    )
@app.get("/download/{filename}")
def download(filename: str):
    safe_filename = Path(filename).name
    file_path = OUTPUT_DIR / safe_filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Plik nie istnieje.")
    return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        filename=safe_filename)