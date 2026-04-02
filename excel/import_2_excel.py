from __future__ import annotations

from dataclasses import asdict, is_dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

INVOICE_COLUMNS = [
    "source_file",
    "template",
    "invoice_number",
    "seller_name",
    "bill_to",
    "ship_to",
    "issue_date",
    "subtotal",
    "discount_amount",
    "shipping",
    "total",
    "items_count",
]

ITEM_COLUMNS = [
    "invoice_number",
    "item_index",
    "description",
    "quantity",
    "rate",
    "amount",
    "category",
]

ERROR_COLUMNS = [
    "source_file",
    "error_type",
    "error_message",
    "template",
]

def export_to_excel(
    invoices: list[Any],
    errors: list[Mapping[str,Any]],
    output_path: str | Path, 
) -> Path:
    """
    Standard ETL export:
      - invoices: 1 wiersz = 1 faktura
      - items:    1 wiersz = 1 pozycja
      - errors:   1 wiersz = 1 błąd (plik/problem)

    Nadpisuje plik output_path.
    """"""
    Standard ETL export:
      - invoices: 1 wiersz = 1 faktura
      - items:    1 wiersz = 1 pozycja
      - errors:   1 wiersz = 1 błąd (plik/problem)

    Nadpisuje plik output_path.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
     #openpyxl tworzy domyślny arkusz — użyjemy go jako "invoices"
    ws_invoices = wb.active
    ws_invoices.title = "invoice"
    ws_items= wb.create_sheet("items")
    ws_errors = wb.create_sheet("errors")

    # --- invoice + items -----
    for inv in invoices:
        inv_dict = _to_dict(inv)

        # invoices row
        inv_row = {
            "source_file": inv_dict.get("source_file"),
            "template": inv_dict.get("template", "unknown"),
            "invoice_number": inv_dict.get("invoice_number"),
            "seller_name": inv_dict.get("seller_name"),
            "bill_to": inv_dict.get("bill_to"),
            "issue_date": _as_excel_date(inv_dict.get("issue_date")),
            "ship_to": inv_dict.get("ship_to"),
            "subtotal": inv_dict.get("subtotal"),
            "discount_amount": inv_dict.get("discount_amount"),
            "shipping": inv_dict.get("shipping"),
            "total": inv_dict.get("total"),
            "items_count": len(inv_dict.get("items") or []),
        }
        _append_row(ws_invoices, INVOICE_COLUMNS, inv_row)

        # items rows
        invoice_number = inv_dict.get("invoice_number")
        items = inv_dict.get("items") or []
        for idx, item in enumerate(items, start=1):
            item_dict = _to_dict(item)
            item_row = {
                "invoice_number": invoice_number,
                "item_index": idx,
                "description": item_dict.get("description"),
                "quantity": item_dict.get("quantity"),
                "rate": item_dict.get("rate"),
                "amount": item_dict.get("amount"),
                "category": item_dict.get("category"),
            }
            _append_row(ws_items, ITEM_COLUMNS, item_row)

    # --- errors ---
    for e in errors:
        #oczekujemy mapping/dict, ale zabezpieczamy sie
        ed = dict(e)
        err_row = {
            "source_file": ed.get("source_file"),
            "error_type": ed.get("error_type"),
            "error_message": ed.get("error_message"),
            "template": ed.get("template", "unknown"),
        }
        _append_row(ws_errors, ERROR_COLUMNS, err_row)
    # opcjonalnie: lekko dopasuj szerokości kolumn (bez stylowania)
    _autosize_columns(ws_invoices, max_rows=300)
    _autosize_columns(ws_items, max_rows=500)
    _autosize_columns(ws_errors, max_rows=300)

    wb.save(output_path)
    return output_path


# ================= helpers =================

def _write_header(ws, columns: list[str]) -> None:
    ws.append(columns)


def _append_row(ws, columns: list[str], row: Mapping[str, Any]) -> None:
    ws.append([row.get(c) for c in columns])

def _to_dict(obj: Any) -> dict[str, Any]:
    """
    Wspiera:
    - pydantic v2: model_dump()
    - pydantic v1: dict()
    - dataclass: asdict()
    - zwykły dict
    - obiekt z __dict__
    """
    if obj is None:
        return {}
    if isinstance(obj, dict):
        return obj

    if is_dataclass(obj):
        return asdict(obj)

    # Pydantic v2
    md = getattr(obj, "model_dump", None)
    if callable(md):
        return md()

    # Pydantic v1
    d = getattr(obj, "dict", None)
    if callable(d):
        return d()

    # fallback
    if hasattr(obj, "__dict__"):
        return dict(obj.__dict__)

    return {}


def _as_excel_date(v: Any) -> Any:
    """
    openpyxl przyjmuje datetime/date i zapisuje jako typ daty w Excelu.
    Jeśli dostaniesz string, zostawiamy string (na MVP).
    """
    if isinstance(v, (date, datetime)):
        return v
    return v


def _autosize_columns(ws, max_rows: int = 300) -> None:
    """
    Lekki autosize bez stylowania.
    Liczymy szerokości na podstawie nagłówka + pierwszych max_rows.
    """
    # openpyxl: ws.iter_rows zwraca komórki
    col_widths: dict[int, int] = {}

    rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows)))
    for r in rows:
        for cell in r:
            if cell.value is None:
                continue
            s = str(cell.value)
            col_idx = cell.column
            col_widths[col_idx] = max(col_widths.get(col_idx, 0), len(s))

    for col_idx, w in col_widths.items():
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(max(w + 2, 10), 60)
